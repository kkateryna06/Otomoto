import requests
import json
from openpyxl import load_workbook
import pandas as pd
import os
import urllib.parse
from datetime import datetime
from bs4 import BeautifulSoup


# Function for loading HTML content
def fetch_html(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36"
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Check for a successful response
        return response.text
    except requests.exceptions.HTTPError as e:
        print(f"Error requesting URL: {url} - {e}")
        return None  # Return None in case of error


# Function to extract JSON data from the first found script
def extract_json_from_html(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    script = soup.find('script', {'type': 'application/ld+json'})
    if script:
        try:
            return json.loads(script.string)
        except json.JSONDecodeError:
            return None
    return None


# Function to extract JSON data from <script id="__NEXT_DATA__"> tag
def extract_json_from_script_tag(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the required tag <script id="__NEXT_DATA__" type="application/json">
    script_tag = soup.find('script', {'id': '__NEXT_DATA__', 'type': 'application/json'})

    if script_tag:
        try:
            return json.loads(script_tag.string)  # Convert the tag content to JSON
        except json.JSONDecodeError:
            print("Error parsing JSON")
            return None
    else:
        print("Tag with required data not found")
        return None


# Function to extract links
def extract_links(html_content):
    # Parsing HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all <a> tags inside <p> tags with class 'e2z61p70 ooa-1ed90th er34gjf0'
    links = []
    p_tags = soup.find_all('p', class_='e2z61p70 ooa-1ed90th er34gjf0')

    for p in p_tags:
        a_tag = p.find('a')  # Search for <a> tag inside <p>

        if a_tag and a_tag.get('href'):  # Check if <a> has href
            links.append(a_tag['href'])  # Add a link to the list

    return links


# Function to convert URL to safe file name
def sanitize_filename(url):
    # Convert the URL to a filesystem-safe string
    return urllib.parse.quote(url, safe='')


# Function to create a safe folder name
def create_safe_folder_name(url):
    folder_name = sanitize_filename(url)
    folder_path = os.path.join("car_photos", folder_name)
    os.makedirs(folder_path, exist_ok=True)  # Create a folder if it doesn't exist
    return folder_path


# Function to upload photos to a specified folder
def download_images(photo_links, folder_path):
    for j, photo_url in enumerate(photo_links):
        try:
            response = requests.get(photo_url, stream=True)
            photo_path = os.path.join(folder_path, f"photo_{j + 1}.jpg")
            with open(photo_path, "wb") as photo_file:
                photo_file.write(response.content)
        except Exception as e:
            print(f"Error uploading photo {photo_url}: {e}")


# Function to clear description from HTML tags
def clean_html_description(description_html):
    # Using BeautifulSoup to Clean HTML
    soup = BeautifulSoup(description_html, "html.parser")
    return soup.get_text(separator=" ").strip()  # Return only text separated by spaces


# Function to retrieve data about the car
def extract_car_data(link, json_data, data):
    advert = json_data.get("props", {}).get("pageProps", {}).get("advert", {})
    # Basic information about the car
    car_name = advert.get("title", "Not specified")
    car_price = advert.get("price", {}).get("value", "Not specified")
    car_description = clean_html_description(advert.get("description", "Description is missing"))
    car_mileage = advert.get("mainFeatures", ["Not specified"])[1] if len(
        advert.get("mainFeatures", [])) > 1 else "Not specified"
    car_mileage = car_mileage[:len(car_mileage) - 3]
    car_fuel_type = advert.get("mainFeatures", ["Not specified"])[3] if len(
        advert.get("mainFeatures", [])) > 3 else "Not specified"
    car_year = advert.get("mainFeatures", ["Not specified"])[0] if len(
        advert.get("mainFeatures", [])) > 0 else "Year not specified"
    car_engine_capacity = advert.get("mainFeatures", ["Not specified"])[2] if len(
        advert.get("mainFeatures", [])) > 2 else "Not specified"

    a = advert.get("details", ["Not specified"])

    body_type = next((i["value"] for i in a if i["key"] == "body_type"), "Not specified")
    gearbox = next((i["value"] for i in a if i["key"] == "gearbox"), "Not specified")
    transmission = next((i["value"] for i in a if i["key"] == "transmission"), "Not specified")

    # Seller information
    seller = advert.get("seller", {})
    seller_type = seller.get("type", "Не указан тип продавца")
    seller_location = seller.get("location", {}).get("shortAddress", "Не указано местоположение")

    # Additional information
    is_accident_free = advert.get("no_accident", "Не указано")

    photos = advert.get("images", {}).get("photos", [])
    photo_links = [photo.get("url") for photo in photos]

    # Specify the base path where the photos will be saved
    base_folder = r"C:\Users\katya\Desktop\otomoto\car_photos"
    photo_folder = sanitize_filename(link)

    # Create paths for photos
    photo_path = os.path.join(base_folder, photo_folder)

    return {
        "Relevant": "yes",
        "Data": data,
        "Sell Data": '',
        "Car Name": car_name,
        "Car Year": int(car_year),
        "Car Fuel Type": car_fuel_type,
        "Car Mileage": int(car_mileage.replace(" ", "")),
        "Car Engine Capacity": car_engine_capacity,
        "Car Price": int(car_price),
        "Body Type": body_type,
        "Gearbox": gearbox,
        "Transmission": transmission,
        "Seller Type": seller_type,
        "Seller Location": seller_location,
        "Car Link": link,
        "Photo Folder": photo_path,
        "Car Description": car_description,
        "Photo Links": photo_links
    }


# The function updates the Excel file while preserving the formats.
def update_excel_with_styles(existing_file, updated_df):
    try:
        # Load an existing Excel file
        workbook = load_workbook(existing_file)
        sheet_name = workbook.sheetnames[0]  # Sheet name (default first)
        sheet = workbook[sheet_name]

        # Clear data but keep styles
        for row in sheet.iter_rows(min_row=2):  # Start with line 2 (the first one is the headings)
            for cell in row:
                cell.value = None  # Delete data, but styles remain

        # Write updated data
        for i, row in updated_df.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=i + 2, column=j + 1, value=value)

        # Save the file
        workbook.save(existing_file)
    except Exception as e:
        print(f"Error updating excel file: {e}")


# Function to retrieve information about machines
def update_data(url):
    data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Load an existing table if it exists
    if os.path.exists("cars_data.xlsx"):
        try:
            existing_df = pd.read_excel("cars_data.xlsx")
            if "Car Link" in existing_df.columns:
                existing_links = set(existing_df["Car Link"].dropna().str.strip())
            else:
                print("The 'Car Link' column is missing from the file.")
                existing_df = pd.DataFrame()
                existing_links = set()
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            existing_df = pd.DataFrame()
            existing_links = set()
    else:
        existing_df = pd.DataFrame()
        existing_links = set()

    # Get the HTML content of the page
    html = fetch_html(url)
    if not html:
        return

    # Extract links to ads
    links = extract_links(html)
    new_car_data = []
    count = 0

    for link in links:
        if link in existing_links:
            print(f"Skip the ad (already in the table): {link}")
            continue

        count+= 1

        # Load the HTML content of the ad
        html = fetch_html(link)
        if not html:
            continue

        print(f"Processing a new ad: {link}")
        json_data = extract_json_from_script_tag(html)
        if json_data:
            car_info = extract_car_data(link, json_data, data)
            new_car_data.append(car_info)

            folder_path = create_safe_folder_name(link)
            download_images(car_info["Photo Links"], folder_path)

        # Define a new column order
        new_column_order = [
            'Relevant', 'Data', 'Sell Data','Car Name', 'Car Year', 'Car Fuel Type', 'Car Mileage', 'Car Engine Capacity', 'Car Price',
            'Body Type', 'Gearbox', 'Transmission', 'Seller Type', 'Seller Location', 'Car Link', 'Photo Folder', 'Car Description'
        ]

        # Save updated data
        if new_car_data:
            new_df = pd.DataFrame(new_car_data)
            updated_df = pd.concat([existing_df, new_df], ignore_index=True)
            updated_df = updated_df[new_column_order]

            # Instead of pandas.to_excel we use openpyxl to save styles
            if os.path.exists("cars_data.xlsx"):
                update_excel_with_styles("cars_data.xlsx", updated_df)
            else:
                updated_df.to_excel("cars_data.xlsx", index=False)  # If the file does not exist, create it
        print(f"Table successfully updated! {count} new ads")

    else:
        print("No new ads found.")


def task():
    url = 'https://www.otomoto.pl/osobowe/mitsubishi/lancer?search%5Bfilter_float_price%3Ato%5D=27000&search%5Border%5D=created_at_first%3Adesc'
    update_data(url)
    with open("logs.txt", "a", encoding="utf-8") as f:  # "a" to append, "w" to overwrite
        f.write(f'The task was completed {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')  # Add line break


# # Task planning
# schedule.every(10).hours.do(task)  # Launch every 10 hours
#
# # Infinite loop for executing tasks
# while True:
#     schedule.run_pending()
#     time.sleep(1)  # Check tasks every second

task()
