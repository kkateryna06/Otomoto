import pandas as pd
import requests
from openpyxl import load_workbook
from datetime import datetime
import os


# Function to check the link
def fetch_html(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36"
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Check for a successful response
        return True
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            return False
        return None


# Function to update the "Relevant" column
def update_relevant_column_with_openpyxl(file_path, date):
    try:
        # Load an existing file
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Default Worksheet

        # Find the indexes of the columns "Car Link" and "Relevant"
        headers = [cell.value for cell in sheet[1]]  # Read the first line (headings)
        car_link_idx = headers.index("Car Link") + 1
        relevant_idx = headers.index("Relevant") + 1
        sell_date_idx = headers.index("Sell Data") + 1

        # Process each line
        for row in range(2, sheet.max_row + 1):
            car_link = sheet.cell(row=row, column=car_link_idx).value
            if car_link:
                print(f"Checking {car_link}")
                is_active = fetch_html(car_link)  # Check if the URL is available
                relevant_value = "yes" if is_active else "no"
                sheet.cell(row=row, column=relevant_idx).value = relevant_value

                if relevant_value == "no":
                    sell_date = datetime.now().strftime(date)
                    sheet.cell(row=row, column=sell_date_idx).value = sell_date

        # Save the file
        workbook.save(file_path)
        print("The file has been successfully updated with formatting preserved!")
    except Exception as e:
        print(f"Error updating Excel file: {e}")


def task():
    file_path = "cars_data.xlsx"
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if os.path.exists(file_path):
        update_relevant_column_with_openpyxl(file_path, date)
        with open("logs.txt", "a", encoding="utf-8") as f:  # "a" to append, "w" to overwrite
            f.write(f'Relevant info was updated {date}\n')  # Лог
    else:
        print("File cars_data.xlsx not found.")


# Starting the task
task()
