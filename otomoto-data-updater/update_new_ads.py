import re

import requests
import json
from openpyxl import load_workbook
import pandas as pd
import os
import urllib.parse
from datetime import datetime
from bs4 import BeautifulSoup

from database_update import update_database, get_all_car_links, get_all_car_links_for_relevant_check


# Function for loading HTML content
def fetch_html(url):
    """
    Loads HTML pages following a URL with a "fake" browser title.
    :param url: URL to load
    :return: HTML content of the page
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36"
    }
    try:
        # Check for a successful response
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.text

    except requests.exceptions.HTTPError as e:
        # Return None in case of error
        print(f"Error requesting URL: {url} - {e}")
        return None


# Function to extract JSON data from the first found script
def extract_json_from_html(html_content):
    """
    Finds a block <script type="application/ld+json"> on the page,
    where there is basic information about the product
    :param html_content: HTML content of the page
    :return: JSON data from the script
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    script = soup.find('script', {'type': 'application/ld+json'})
    if script:
        try:
            return json.loads(script.string)
        except json.JSONDecodeError:
            return None
    return None


# Function to extract JSON data from <script id="__NEXT_DATA__"> tag
def extract_json_from_script_tag(html_content):
    """
    The main function is to look for the
    <script id="__NEXT_DATA__" type="application/json"> tag,
    which contains full information about the ad. This is the main source of data.
    :param html_content: HTML content of the page
    :return: JSON data from the script tag
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the required tag <script id="__NEXT_DATA__" type="application/json">
    script_tag = soup.find('script', {'id': '__NEXT_DATA__', 'type': 'application/json'})

    if script_tag:
        try:
            return json.loads(script_tag.string)  # Convert the tag content to JSON
        except json.JSONDecodeError:
            print("Error parsing JSON")
            return None
    else:
        print("Tag with required data not found")
        return None


# Function to extract links
def extract_links(html_content):
    """
    Finds all ad links that contain otomoto.pl and ID in the URL. Returns set() of unique links.
    :param html_content: HTML content of the page
    :return: set() of unique links
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all <a> tags containing links to ads
    links = set()
    for a_tag in soup.find_all('a', href=True):
        href = a_tag['href']
        if "otomoto.pl" in href and "ID" in href:
            links.add(href)

    return links


# Function to convert URL to safe file name
def sanitize_filename(url):
    """
    Converts URLs to safe file names.
    Used to store photos and HTML to avoid characters that are not allowed in file names.
    :param url: URL to convert
    :return: Safe file name
    """
    return urllib.parse.quote(url, safe='')


def create_safe_folder_name(url):
    """
    Converts URLs to safe folder names.
    Used to store photos and HTML to avoid characters that are not allowed in folder names.
    :param url: URL to convert
    :return: Safe folder name
    """
    folder_name = sanitize_filename(url)
    folder_path = os.path.join("car_photos", folder_name)

    # Create a folder if it doesn't exist
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def extract_number(value):
    """
    Extracts a number from a string. Eg: "180 KM" → 180.0
    :param value: String to extract number from
    :return: Extracted number or None if not found
    """
    match = re.search(r'\d+(\.\d+)?', value)
    return float(match.group()) if match else None


# Function to upload photos to a specified folder
def download_images(photo_links, folder_path):
    """
    Downloads photos from links and saves them in the car_photos folder.
    :param photo_links: List of photo URLs
    :param folder_path: Path to the folder where photos will be saved
    """
    for j, photo_url in enumerate(photo_links):
        try:
            response = requests.get(photo_url, stream=True)
            photo_path = os.path.join(folder_path, f"photo_{j + 1}.jpg")
            with open(photo_path, "wb") as photo_file:
                photo_file.write(response.content)
        except Exception as e:
            print(f"Error uploading photo {photo_url}: {e}")


# Function to clear description from HTML tags
def clean_html_description(description_html):
    """
    Cleans the description from HTML tags - extracts clean text.
    :param description_html: HTML description
    :return: Clean text from the description
    """
    # Using BeautifulSoup to Clean HTML
    soup = BeautifulSoup(description_html, "html.parser")

    # Return only text separated by spaces
    return soup.get_text(separator=" ").strip()

def extract_param(label, parameters, is_label=False):
    """
    Helper function for safely retrieving values from declaration JSON parameters.
    Works with both .label and .value.
    :param label: Parameter label
    :param parameters: Declaration JSON parameters
    :param is_label: Whether the parameter is a label or not
    :return: Parameter value or None if not found
    """
    try:
        if is_label:
            return parameters.get(label, {}).get("values")[0].get("label")
        else:
            return parameters.get(label, {}).get("values")[0].get("value")
    except:
        return None

# Function to retrieve data about the car
def extract_car_data(link, json_data):
    """
    Key function. Receives JSON ads and extracts from it:
        - car parameters (brand, model, fuel type, body, gearbox, mileage, etc.),
        - price, date, description, seller,
        - photos and location,
        - saves HTML and photos to local folders,
        - returns a dictionary with all this info.
    :param link: Link to the ad
    :param json_data: JSON data from the script tag
    :return: Dictionary with car data
    """
    advert = json_data.get("props", {}).get("pageProps", {}).get("advert", {})

    # BASIC INFORMATION
    parameters = advert.get("parametersDict", {})

    try:
        mark = parameters.get("make", {}).get("values")[0].get("label")
        model = extract_param("model", parameters, True)
        version = extract_param("version", parameters, True)
        color = extract_param("color", parameters)
        door_count = extract_param("door_count", parameters)
        nr_seats = extract_param("nr_seats", parameters)
        year = extract_param("year", parameters)
        generation = extract_param("generation", parameters, True)

        # TECHNICAL SPECS
        fuel_type = extract_param("fuel_type", parameters, True)
        engine_capacity = extract_param("engine_capacity", parameters)
        engine_power = extract_param("engine_power", parameters)
        body_type = extract_param("body_type", parameters, True)
        gearbox = extract_param("gearbox", parameters, True)
        transmission = extract_param("transmission", parameters, True)
        extra_urban_consumption = extract_param("extra_urban_consumption", parameters)
        urban_consumption = extract_param("urban_consumption", parameters)
        mileage = extract_param("mileage", parameters)

        # CONDITION HISTORY
        if extract_param("registered", parameters) == "1":
            has_registration = True
        else:
            has_registration = False

        # EQUIPMENT
        equipment = advert.get("equipment", {})
        equipment = json.dumps(equipment)

        # PARAMETERS
        parameters_dict = advert.get("parametersDict", {})
        parameters_dict = json.dumps(parameters_dict)

        # ADVERT INFO
        price = advert.get("price", {}).get("value", None)
        date = advert.get("createdAt", None)
        id = advert.get("id")
        description = clean_html_description(advert.get("description", None))

        seller = advert.get("seller", {})
        seller_type = seller.get("type", None)

        # LOCATION
        location = advert.get("seller", {}).get("location", {}).get("map", {})
        location = json.dumps(location)

        # PATH
        base_folder_photo = r"C:\Users\katya\Desktop\otomoto\otomoto-data-updater\car_photos"
        photo_folder = urllib.parse.quote(link, safe='')
        photo_path = os.path.join(base_folder_photo, photo_folder)

        base_folder_html = r"C:\Users\katya\Desktop\otomoto\otomoto-data-updater\car_htmls"
        html_folder = urllib.parse.quote(link, safe='')
        html_path = os.path.join(base_folder_html, html_folder)

        # Download photos
        photos = advert.get("images", {}).get("photos", [])
        photo_links = [photo.get("url") for photo in photos]
        folder_path = create_safe_folder_name(link)
        download_images(photo_links, folder_path)

        dict_result = {
            "mark": mark, "model": model, "version": version, "color": color,
            "door_count": door_count, "nr_seats": nr_seats, "year": year, "generation": generation,
            "fuel_type": fuel_type, "engine_capacity": engine_capacity, "engine_power": engine_power,
            "body_type": body_type, "gearbox": gearbox,
            "transmission": transmission, "urban_consumption": urban_consumption,
            "extra_urban_consumption": extra_urban_consumption, "mileage": mileage,
            "has_registration": has_registration, "equipment": equipment,
            "parameters_dict": parameters_dict, "price": price, "date": date, "description": description,
            "link": link, "car_id": id, "location": location, "photo_path": photo_path,
            "html_path": html_path, "seller_type": seller_type
        }
        return dict_result
    except Exception as e:
        print(f"Error: {e}")
        return None


# The function updates the Excel file while preserving the formats.
def update_excel_with_styles(existing_file, updated_df):
    """
    Clears the contents of the Excel file, but preserves styles. Then writes the updated data.
    :param existing_file: Path to the existing Excel file
    :param updated_df: Updated DataFrame with data to write
    """
    try:
        # Load an existing Excel file
        workbook = load_workbook(existing_file)
        sheet_name = workbook.sheetnames[0]  # Sheet name (default first)
        sheet = workbook[sheet_name]

        # Clear data but keep styles
        for row in sheet.iter_rows(min_row=2):  # Start with line 2 (the first one is the headings)
            for cell in row:
                cell.value = None  # Delete data, but styles remain

        # Write updated data
        for i, row in updated_df.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=i + 2, column=j + 1, value=value)

        # Save the file
        workbook.save(existing_file)
    except Exception as e:
        print(f"Error updating excel file: {e}")


# Function to retrieve information about machines
def update_data(url, database_table, excel_table):
    """
    Step by step:
        - Loads existing ad links from the DB.
        - Parses HTML by the given url, finds all new ad links.
        - For each new ad:
        - Loads HTML.
        - Saves HTML to car_htmls.
        - Parses JSON from the script.
        - Extracts data via extract_car_data.
        - Updates the database (update_database).
        - Prepares data for Excel.
    :param url: URL to parse
    :param database_table: Name of the database table to update
    :param excel_table: Name of the Excel file to update
    """
    # Load an existing table if it exists
    existing_links = get_all_car_links(database_table)

    # Get the HTML content of the page
    html = fetch_html(url)
    if not html:
        return

    # Extract links to ads
    links = extract_links(html)
    new_car_data = []
    count = 0

    for link in links:
        if link in existing_links:
            continue

        count+= 1

        # Load the HTML content of the ad
        html = fetch_html(link)
        if not html:
            continue


        html_filename = sanitize_filename(link) + ".html"
        html_path = os.path.join('car_htmls', html_filename)

        os.makedirs('car_htmls', exist_ok=True)

        with open(html_path, "w", encoding="utf-8") as file:
            file.write(html)

        print(f"Processing a new ad: {link}")
        json_data = extract_json_from_script_tag(html)
        if json_data:
            # Define a new column order
            new_column_order = [
                'relevant', 'date', 'sell_date', 'mark', 'model', 'version', 'year', 'fuel_type', 'mileage',
                'engine_capacity', 'price', 'body_type', 'gearbox', 'transmission',
                'seller_type', 'location', 'link', 'photo_path', 'html_path', 'description'
            ]

            car_info = extract_car_data(link, json_data)
            if car_info is not None:
                update_database(car_info, "new_ads", database_table)
                filtered_dict = {key: car_info[key] for key in car_info if key in new_column_order}
                try: filtered_dict["year"] = int(float(filtered_dict["year"]))
                except: pass
                filtered_dict["price"] = int(float(filtered_dict["price"]))
                try: filtered_dict["mileage"] = int(filtered_dict["mileage"])
                except: pass
                try: filtered_dict["engine_capacity"] = filtered_dict["engine_capacity"] + " cm3"
                except: pass
                filtered_dict["relevant"] = "yes"
                new_car_data.append(filtered_dict)

    # Save updated date
    if new_car_data:
        print(f"Table successfully updated! {count} new ads")

    else:
        print("No new ads found.")


def load_links_from_file(filepath="links_config.txt"):
    all_links = []
    special_links = []

    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()

    current_type = None
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith("#"):
            if "ALL" in line.upper():
                current_type = "all"
            elif "SPECIAL" in line.upper():
                current_type = "special"
            continue
        if current_type == "all":
            all_links.append((line, None))  # no label
        elif current_type == "special":
            label = None
            if "|" in line:
                url, label = line.split("|", 1)
                line = url.strip()
                label = label.strip()
            special_links.append((line, label or "Special ad"))
    return all_links, special_links


def run_links(links, db_table, excel_table):
    for url, label in links:
        update_data(url, db_table, excel_table)
        with open("logs.txt", "a", encoding="utf-8") as f:
            f.write(f'{label or url} was updated {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
